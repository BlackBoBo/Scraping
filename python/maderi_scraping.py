import os
import time
import shutil
import random
from pprint import pprint
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, PatternFill
import pendulum
import itertools
import subprocess
import module.util as util
from selenium import webdriver
from datetime import datetime, timedelta
from module.slack_alert import SlackAlert

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException, NoSuchElementException

from airflow.decorators import dag, task
from airflow.hooks.base_hook import BaseHook
from airflow.providers.microsoft.mssql.hooks.mssql import MsSqlHook
from airflow.providers.postgres.hooks.postgres import PostgresHook

local_tz = pendulum.timezone("Asia/Seoul")
alter = SlackAlert('airflow')

mssql = MsSqlHook(mssql_conn_id="")

default_args = {
    'owner': 'yeongjin',
    'start_date': datetime(2023, 6, 12, tzinfo=local_tz),
    'provide_context': False,
    # 'on_success_callback': alter.slack_data_report,
    "on_failure_callback": alter.slack_fail_alert
}

maderi_info = BaseHook.get_connection('')
maderi_id = maderi_info.login
maderi_pw = maderi_info.password

@dag(
    # schedule_interval=None,
    schedule_interval="30 12 * * *",
    default_args=default_args,
    # max_active_runs=1,
    tags=['daily', '', "crawling"]
)


def maderi_crawling():
    """
        registered_keyword_total_data : 등록된 키워드의 전체 데이터 excel 파일을 다운로드 후 s3 업로드
        searched_keyword_total_data : 키워드를 받고 전체 데이터 excel 파일을 다운로드 후 s3 업로드
        related_keyword_total_data : 등록된 키워드의 연관검색어 excel 파일을 다운로드 후, 기존 excel 파일에 데이터 최신화 후 s3 업로드
    """

    def is_leap_year(year):  # 윤년 true
        return (year % 4 == 0) and (year % 100 != 0 or year % 400 == 0)

    def get_week(year, mon, day):
        if mon <= 2:
            year -= 1
            mon += 12
        a = year // 100  # 년도의 앞2자리
        b = year % 100  # 년도의 뒤2자리
        week = ((21 * a // 4) + (5 * b // 4) + 26 * (mon + 1) // 10 + day - 1) % 7
        return week

    def find_number_indices(lst, number):
        indices = []
        for i in range(len(lst)):
            for j in range(len(lst[i])):
                if lst[i][j] == number:
                    indices.append((i + 1, j + 1))
        return indices[0]

    def get_calendar(year, month, day):
        year = int(year)
        month = int(month)
        day = int(day)

        last_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        last_day = last_days[month - 1] if not (is_leap_year(year) and month == 2) else 29  # 윤년

        #       일,월,화,수,목,금,토
        cal_list = [[0, 0, 0, 0, 0, 0, 0]]
        cal_list += [[0, 0, 0, 0, 0, 0, 0]]
        cal_list += [[0, 0, 0, 0, 0, 0, 0]]
        cal_list += [[0, 0, 0, 0, 0, 0, 0]]
        cal_list += [[0, 0, 0, 0, 0, 0, 0]]
        cal_list += [[0, 0, 0, 0, 0, 0, 0]]

        col = get_week(year, month, 1)
        row = 0

        for day2 in range(1, last_day + 1):
            cal_list[row][col] = day2
            col += 1
            if col > 6:  # 다음 행에다 날짜기입
                col = 0
                row += 1

        indices = find_number_indices(cal_list, day)
        return indices

    @task
    def registered_keyword_total_data():
        """
            사이트 접속 -> 로그인 -> 키워드 목록 받기 -> (반복지점) 키워드 선택(최대 10개) -> 키워드 분석 -> 16.01.01 ~ 오늘 설정 -> 키워드별 통계 다운로드 -> s3 저장-> (반복)
        """

        def selenium_execute(target, path, sec=0):
            try:
                if target == 'css':
                    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                    element.click()

                elif target == 'xpath':
                    element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                    element.click()

            except ElementClickInterceptedException:
                print("ElementClickInterceptedException error occurred, [Retrying..]")
                driver.refresh()
                try:
                    if target == 'css':
                        element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                        element.click()

                    elif target == 'xpath':
                        element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                        element.click()

                except Exception as e:
                    err_msg = "selenium_execute ElementClickInterceptedException Error : " + str(e)
                    print(err_msg)
                    # raise err_msg
            except TimeoutException:
                print("TimeoutException error occurred, [Retrying..]")
                driver.refresh()
                try:
                    if target == 'css':
                        element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                        element.click()

                    elif target == 'xpath':
                        element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                        element.click()
                except Exception as e:
                    err_msg = "selenium_execute TimeoutException Error : " + str(e)
                    raise err_msg

            # time.sleep(random.randint(3, 7))  # 실전용

        def change_folder_name(old_name, new_name, add_dir_path=''):
            old_filename = os.path.join(downloads_path, old_name)  # 다운로드한 기존 파일명을
            new_filename = os.path.join(downloads_path + add_dir_path, new_name)  # 원하는 파일명으로 변경
            shutil.move(old_filename, new_filename)

        data = {}
        date = util.get_date().replace("-","")  # '어제' 데이터가 네이버에 의존하며, 업데이트 시간이 오전 10~12시
        maderi_bucket = ""

        # 경로 설정
        cat = "keyword_trend_daily"
        present_path = os.getcwd()
        downloads_path = os.path.join(present_path, "data", cat)
        print(present_path)
        print(downloads_path)

        # 드라이버 옵션 설정
        chrome_options = Options()
        prefs = {
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)

        chrome_options.add_experimental_option("prefs", {"download.default_directory": f"{downloads_path}"})
        chrome_options.add_argument("--headless")  # 서버용
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")  # 서버용
        # chrome_options.add_argument("--disable-gpu")
        # chrome_options.add_argument("--blink-settings=imagesEnabled=false")  # 이미지 로딩 false
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")  # window 10, chrome

        # 드라이버 설정
        driver_path = ChromeDriverManager().install()  # 드라이버 설치
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 100)
        driver.set_window_size(1920, 1080)

        # 키워드의 최근 업데이트 날짜 가져오기
        sql = """"""
        temp_keyword_last_updated_date = {}
        keyword_last_updated_date = util.query_ods("select", sql)
        keyword_last_updated_date = [[str(values[0]), values[1]] for values in keyword_last_updated_date]
        for sublist in keyword_last_updated_date:
            key = sublist[1]
            value = sublist[0].replace(" 00:00:00", "")

            temp_keyword_last_updated_date[key] = value

        try:
            # 마대리 접속
            url = 'https://www.maderi.co.kr'  # 홈 ( 로그인 )
            driver.get((url + "/login"))
            driver.implicitly_wait(100)  # load

            # 로그인 정보 입력 및 로그인
            id_xpath = "/html/body/div[1]/section/div[1]/div/div/form/div[1]/div/input"
            passwd_xpath = "/html/body/div[1]/section/div[1]/div/div/form/div[2]/div/input"
            login_btn_xpath = "/html/body/div[1]/section/div[1]/div/div/form/button"

            driver.find_element(By.ID, 'email').send_keys(maderi_id)
            driver.find_element(By.ID, 'password').send_keys(maderi_pw)
            selenium_execute('xpath', login_btn_xpath)

            # 로그인 대기
            time.sleep(30)

            # keyword 페이지 이동
            driver.get((url + "/keyword"))

            # 키워드 목록 파일 다운로드
            set_calendar_open_xpath = "/html/body/div[1]/main/div[2]/div/div/div/input"
            selenium_execute("xpath", set_calendar_open_xpath)
            set_date_yesterday_keyword_list_xpath = "/html/body/div[1]/main/div[2]/div/div/div/div/div/div/button[1]"
            selenium_execute("xpath", set_date_yesterday_keyword_list_xpath)

            keyword_excel_download_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[2]/div[1]/div[2]/button[3]"
            selenium_execute("xpath", keyword_excel_download_btn_xpath, 7)
            time.sleep(7)  # 파일 다운로드 대기

            desired_filename = f"Keyword_키워드_종합_진단_{date}.xlsx"  # 저장할 파일명 셋팅
            change_folder_name(r"Keyword_키워드 종합 진단.xlsx", desired_filename)

            # 전체 키워드 가져오기 및 s3 업로드
            if os.path.isfile(os.path.join(downloads_path, desired_filename)):
                workbook = load_workbook(filename=os.path.join(downloads_path, desired_filename))
                sheet = workbook.active

                keyword_list = []
                start_cell = 'C7'
                end_cell = 'C106'
                for row in sheet[start_cell:end_cell]:
                    for cell in row:
                        keyword_list.append(cell.value)

                keyword_list = list(filter(None, keyword_list))
                keyword_cnt = len(keyword_list)  # None 제거 후 개수 카운트
                print(keyword_list)

                try:
                    save_command_keyword_file_s3 = f"aws s3 cp {downloads_path}/{desired_filename} '{maderi_bucket}/키워드리스트/{date[:4]}/{date[4:6]}/{desired_filename}'"
                    result = subprocess.run(save_command_keyword_file_s3, shell=True, capture_output=True, text=True)
                except Exception as e:
                    raise e

                if result.returncode == 0:
                    print('키워드 파일 업로드 성공')
                    print('표준 출력:', result.stdout)
                else:
                    print('키워드 파일 업로드 실패')
                    print('에러 출력:', result.stderr)

                os.system(f"rm {downloads_path}/{desired_filename}")
                print("키워드 파일 삭제완료")

                time.sleep(10)
            else:
                raise Exception("Keyword_File Is Not Exist")

            # 순차적 검색
            start_keyword = 1 # 오류 발생 시, 출력된 keyword_cnt : 값을 넣으면 실패한 지점부터 실행
            last_cnt = start_keyword - 1
            for cnt in range(start_keyword, keyword_cnt + 1):
                # 현재 목록이 100개로 설정된 경우 스킵
                cnt_xpath = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/input"
                value = driver.find_element(By.XPATH, cnt_xpath).get_attribute("value")
                if value != '100개':
                    # 100개 보기
                    open_page_option_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/input"
                    driver.find_element(By.XPATH, open_page_option_btn_xpath).click()

                    # 100개 보기 선택
                    option_100 = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/ul/li[4]"
                    selenium_execute('xpath', option_100, 3)

                    # print(data) # 데이터 확인용
                print("keyword_cnt : ", cnt)

                # 분석할 키워드 선택
                select_box_xpath = f"/html/body/div[1]/main/div[4]/article/div/div[2]/div[2]/table/tbody/tr[{cnt}]/td[1]/div/label"
                selenium_execute('xpath', select_box_xpath)

                if (cnt % 10 == 0) or (cnt == keyword_cnt):  # 10개 or 마지막
                    # 검색분석 클릭
                    analysis_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[3]/button[2]"
                    selenium_execute('xpath', analysis_btn_xpath, 10)

                    # 일별 보기 클릭
                    daily_setting_btn_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[1]/button[2]"
                    selenium_execute('xpath', daily_setting_btn_xpath, 10)
                    time.sleep(5)

                    # 캘린더 열기
                    open_set_yesterday_option_btn_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/input"
                    selenium_execute('xpath', open_set_yesterday_option_btn_xpath, 5)

                    # 전체데이터 다운을 위한 날짜선택
                    select_year_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[2]/button"
                    select_2016_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[2]/ul/li[1]"
                    selenium_execute("xpath", select_year_xpath)
                    selenium_execute("xpath", select_2016_xpath)

                    select_month_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[1]/button"
                    select_01_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[1]/ul/li[1]"
                    selenium_execute("xpath", select_month_xpath)
                    selenium_execute("xpath", select_01_xpath)

                    select_16_06_01_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/tbody/tr[1]/td[6]/button"
                    selenium_execute("xpath", select_16_06_01_xpath)

                    col, row = get_calendar(date[:4], date[4:6], date[6:])
                    select_today_xpath = f"/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[2]/table/tbody/tr[{col}]/td[{row}]/button"
                    selenium_execute("xpath", select_today_xpath)

                    # 날짜 선택 후 적용 버튼
                    temp_selectbtn = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[2]/button[2]"
                    selenium_execute('xpath', temp_selectbtn, 2)

                    # pc, mobile 포함 체크
                    include_pc_btn_xpath = "/html/body/div[1]/main/div[5]/div[3]/div[3]/label[1]"
                    selenium_execute('xpath', include_pc_btn_xpath, 1)
                    time.sleep(1)
                    include_mobile_btn_xpath = "/html/body/div[1]/main/div[5]/div[3]/div[5]/label[1]"
                    selenium_execute('xpath', include_mobile_btn_xpath, 1)

                    # 선택된 키워드 돌리기
                    for idx, _ in enumerate(keyword_list[last_cnt:cnt], start=1):
                        # 다음 키워드 선택
                        keyword_xpath = f"/html/body/div[1]/main/div[4]/div[{idx}]"
                        selenium_execute('xpath', keyword_xpath, 30)
                        time.sleep(7)

                        # 검색량추이 엑셀 다운로드
                        search_cnt_trend_download_btn_xpath = "/html/body/div[1]/main/div[5]/div[1]/div[2]/button"
                        selenium_execute('xpath', search_cnt_trend_download_btn_xpath, 7)
                        time.sleep(15)

                        # 파일 읽어서 키워드 추출, 데이터 저장
                        if os.path.isfile(os.path.join(downloads_path, "Keyword_검색분석_검색량 트렌드.xlsx")):  # 다운로드 되는 파일명
                            workbook = load_workbook(filename=os.path.join(downloads_path, "Keyword_검색분석_검색량 트렌드.xlsx"))
                            sheet = workbook.active

                            last_column_value = len(sheet["A"])
                            keyword = sheet["B2"].value

                            if keyword not in temp_keyword_last_updated_date.keys(): # 엑셀의 키워드가 db에 없는 단어면,
                                temp_keyword_last_updated_date[keyword] = '2015-12-31'

                            # 특정 셀 영역 데이터 가져온 후, 이중 리스트 언패킹, 클래스 속성 값을 리스트에 저장
                            date_column = [obj.value for obj in list(itertools.chain(*sheet["A8":f"A{last_column_value}"]))]
                            total_column = [obj.value for obj in list(itertools.chain(*sheet["B8":f"B{last_column_value}"]))]
                            pc_column = [obj.value for obj in list(itertools.chain(*sheet["C8":f"C{last_column_value}"]))]
                            mobile_column = [obj.value for obj in list(itertools.chain(*sheet["D8":f"D{last_column_value}"]))]

                            data[keyword] = {}  # 초기화
                            for idx2, maderi_date in enumerate(date_column):
                                if maderi_date > temp_keyword_last_updated_date[keyword]: # date_column 값이 키워드 업데이트 날짜보다 큰 경우에만 인서트한다.
                                    data[keyword][maderi_date] = {}
                                    data[keyword][maderi_date].update({
                                        "total_cnt": total_column[idx2],
                                        "pc_cnt": pc_column[idx2],
                                        "mobile_cnt": mobile_column[idx2]
                                    })

                            keyword_latest_updated_date = temp_keyword_last_updated_date[keyword].replace('-', '')  # 최근에 업데이트 된 날짜의 파일

                            delete_filename = f'{keyword}_검색분석_검색량_트렌드_{keyword_latest_updated_date}.xlsx'
                            desired_filename = f"{keyword}_검색분석_검색량_트렌드_{date}.xlsx"  # 저장할 파일명 셋팅

                            # s3에 저장
                            try:
                                # 저장 전 삭제
                                # 기존 분리한 파일 삭제
                                delete_keyword_file_before_download = f"aws s3 rm '{maderi_bucket}/키워드검색량/{delete_filename}'"
                                result = subprocess.run(delete_keyword_file_before_download, shell=True, capture_output=True, text=True)
                                if result.returncode == 0:
                                    print('검색량 파일 삭제 성공')
                                    print('표준 출력:', result.stdout)
                                else:
                                    print('검색량 삭제 실패')
                                    print('에러 출력:', result.stderr)

                                save_command_keyword_file_s3 = f"aws s3 cp {downloads_path}/* 's3://edl-prod-archivei-bucket/maderi/키워드검색량/{desired_filename}'"
                                result = subprocess.run(save_command_keyword_file_s3, shell=True, capture_output=True, text=True)
                                if result.returncode == 0:
                                    print('검색량 파일 업로드 성공')
                                    print('표준 출력:', result.stdout)
                                else:
                                    print('검색량 파일 업로드 실패')
                                    print('에러 출력:', result.stderr)

                                os.system(f"rm {downloads_path}/*")
                            except Exception as err:
                                raise err

                        else:
                            print(keyword_list[cnt-1] + " Trend File Not Exists")  # 키워드 나오도록 수정필요

                    last_cnt = cnt
                    driver.get(url + "/keyword")

            print(data)

            # 드라이버 종료
            driver.quit()

            print(temp_keyword_last_updated_date)

            try:
                # DB 인서트
                insert_query = f"""INSERT INTO DATA_DW.dbo.DW_KEYWORD_STAT(KS_DATE, KS_KEYWORD, KS_TOT_CNT, KS_PC_CNT, KS_MOBILE_CNT, KS_UPD_DT) VALUES(%s, %s, %d, %d, %d, GETDATE())"""
                insert_sql_args = []

                for keyword, db_date in temp_keyword_last_updated_date.items():
                    if keyword in data.keys():
                        for maderi_date in data[keyword]:
                            if maderi_date > db_date:
                                insert_sql_args.append((maderi_date,
                                                        keyword,
                                                        data[keyword][maderi_date]['total_cnt'],
                                                        data[keyword][maderi_date]['pc_cnt'],
                                                        data[keyword][maderi_date]['mobile_cnt']))

                insert_sql_args = list(set(insert_sql_args)) # 중복제거
                print(insert_sql_args)
                util.query_ods("insert", insert_query, insert_sql_args)
            except Exception as e:
                print("DB insert Error")
                raise e
        except Exception as e:
            print("Collecting Error")
            driver.quit()
            raise e
        os.system(f"rm {downloads_path}/*")

    @task
    def searched_keyword_total_data():
        # -*- coding: utf-8 -*-
        print(datetime.now())
        # 사이트 접속 -> 로그인 -> 키워드 목록 받기 -> (반복지점) 키워드 정렬 -> 키워드 선택(최대) -> 키워드 분석 -> '어제'로 날짜 설정 -> 키워드별 통계 다운로드 -> (반복)

        def selenium_execute(target, path, sec=0):
            # print(path)
            try:
                if target == 'css':
                    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                    element.click()

                elif target == 'xpath':
                    element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                    element.click()

            except ElementClickInterceptedException:
                driver.refresh()

                if driver.find_element(By.XPATH, path).is_displayed():
                    try:
                        if target == 'css':
                            element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                            element.click()

                        elif target == 'xpath':
                            element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                            element.click()

                    except Exception as e:
                        err_msg = "selenium_execute Error : " + str(e)
                        raise err_msg
            # time.sleep(random.randint(3, 7))  # 실전용

        def split_searching_analysis_data_and_upload_s3_and(kywrd, keyword_latest_updated_date, keyword_total_cnt, keyword_pc_cnt, keyword_mob_cnt):
            print(kywrd, "데이터 분리 및 업데이트 시작")
            new_workbook = Workbook()
            new_sheet = new_workbook.active

            fill = PatternFill(fill_type='solid', fgColor='DDEBF7')
            border_style = Side(border_style="thin", color="000000")  # 얇은 실선, 검은색
            border = Border(top=border_style, right=border_style, bottom=border_style, left=border_style)

            new_sheet["A2"].font = Font(bold=True)
            new_sheet["A3"].font = Font(bold=True)
            new_sheet["A6"].font = Font(bold=True)

            new_sheet["B6"].font = Font(bold=True)
            new_sheet["B7"].font = Font(bold=True)
            new_sheet["C7"].font = Font(bold=True)
            new_sheet["D7"].font = Font(bold=True)

            new_sheet["A6"].fill = fill
            new_sheet["B6"].fill = fill
            new_sheet["B7"].fill = fill
            new_sheet["C7"].fill = fill
            new_sheet["D7"].fill = fill

            new_sheet["A6"].border = border
            new_sheet["B6"].border = border
            new_sheet["B7"].border = border
            new_sheet["C7"].border = border
            new_sheet["D7"].border = border

            new_sheet["A2"] = "분석 대상"
            new_sheet["A3"] = "분석 기간"
            new_sheet["A6"] = "날짜"

            new_sheet.merge_cells("A6:A7")
            new_sheet.merge_cells("B6:D6")

            new_sheet["B3"] = f"{date_column[0]} ~ " + date[:4] + "-" + date[4:6] + "-" + date[6:]
            new_sheet["B6"] = kywrd

            new_sheet["B7"] = "총합"
            new_sheet["C7"] = "PC"
            new_sheet["D7"] = "모바일"

            new_sheet["B2"] = kywrd

            for idx, excel_date in enumerate(date_column):
                new_sheet[f"A{idx + 8}"] = excel_date
                new_sheet[f"B{idx + 8}"] = keyword_total_cnt[idx]
                new_sheet[f"C{idx + 8}"] = keyword_pc_cnt[idx]
                new_sheet[f"D{idx + 8}"] = keyword_mob_cnt[idx]

                new_sheet[f"A{idx + 8}"].border = border
                new_sheet[f"B{idx + 8}"].border = border
                new_sheet[f"C{idx + 8}"].border = border
                new_sheet[f"D{idx + 8}"].border = border

            # 업로드 전 삭제
            keyword_latest_updated_date = keyword_latest_updated_date.replace('-', '') # 최근에 업데이트 된 날짜의 파일

            filename = f'{kywrd}_검색분석_검색량_트렌드_{date}.xlsx'
            delete_filename = f'{kywrd}_검색분석_검색량_트렌드_{keyword_latest_updated_date}.xlsx'

            if kywrd is not None:
                new_workbook.save(f"{downloads_path}/{filename}")

                # 기존 분리한 파일 삭제
                delete_keyword_file_before_download = f"aws s3 rm '{maderi_bucket}/키워드검색량/{delete_filename}'"
                result = subprocess.run(delete_keyword_file_before_download, shell=True, capture_output=True, text=True)
                if result.returncode == 0:
                    print(f'{kywrd} 파일 삭제 성공')
                    print('표준 출력:', result.stdout)
                else:
                    print(f'{kywrd} 파일 삭제 실패')
                    print('에러 출력:', result.stderr)

                # 분리한 파일 업로드
                save_keyword_file_s3 = f"aws s3 cp {downloads_path}/{filename} '{maderi_bucket}/키워드검색량/{filename}'"
                result = subprocess.run(save_keyword_file_s3, shell=True, capture_output=True, text=True)

                if result.returncode == 0:
                    print(f'{kywrd} 파일 업로드 성공')
                    print('표준 출력:', result.stdout)
                else:
                    print(f'{kywrd} 파일 업로드 실패')
                    print('에러 출력:', result.stderr)

                os.system(f"rm {downloads_path}/{filename}")

        # 검색할 키워드 리스트 정보q
        data = {}
        dict_keywords = {}
        date = util.get_date().replace("-","")  # '어제' 데이터가 네이버에 의존하며, 업데이트 시간이 오전 10~12시
        maderi_bucket = ""

        # 경로 설정
        cat = "search_trend_daily"
        present_path = os.getcwd()
        downloads_path = os.path.join(present_path, f"data", cat)

        # 드라이버 옵션 설정
        chrome_options = Options()
        chrome_options.add_experimental_option("prefs", {"download.default_directory": f"{downloads_path}"})
        chrome_options.add_argument("--headless")  # 서버용
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")  # 서버용
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")  # window 10, chrome

        # 드라이버 설정
        driver_path = ChromeDriverManager().install()  # 드라이버 설치
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 100)
        driver.set_window_size(1920, 1080)

        # 키워드 최신 업데이트 날짜 select
        sql = """"""
        temp_keyword_last_updated_date = {}
        keyword_last_updated_date = util.query_ods("select", sql)
        keyword_last_updated_date = [[str(values[0]), values[1]] for values in keyword_last_updated_date]
        for row in keyword_last_updated_date:  # [0] : 업데이트 날짜, [1] : 키워드
            key = row[1]
            value = row[0].replace(" 00:00:00", "")
            temp_keyword_last_updated_date[key] = value

        # 키워드 정보 가져오기
        sql = ""
        db_keyword_result = util.query_ods("select", sql)
        db_keyword_result.sort()
        # db_keyword_result = db_keyword_result[80:]  # 일정 범위만 테스트 or 실패 했을 때 사용
        print("db_keyword_result value : ", db_keyword_result)

        # 키워드, 유의어를 dict 타입으로 정리
        for row in db_keyword_result:
            dict_keywords[row[0]] = {}
            dict_keywords[row[0]] = row[1].split(',')  # [0]: 키워드, [1] : 유사어


        driver.implicitly_wait(100)  # load
        url = 'https://www.maderi.co.kr'  # 홈 ( 로그인 )
        driver.get((url + "/login"))

        # 로그인 정보 입력 및 로그인
        login_btn_xpath = "/html/body/div[1]/section/div[1]/div/div/form/button"
        driver.find_element(By.ID, 'email').send_keys(maderi_id)
        driver.find_element(By.ID, 'password').send_keys(maderi_pw)
        selenium_execute('xpath', login_btn_xpath)

        # 로그인 대기
        time.sleep(5)

        # 마대리 접속
        url = 'https://www.maderi.co.kr'  # 홈 ( 로그인 )
        driver.get((url + "/keyword/simple"))

        keywords = dict_keywords
        adjust_idx_value = 0  # idx 값을 재활용해서 유사어 넣을 사d용하는 값
        # 키워드 입력
        for idx, keyword in enumerate(keywords, start=1):
            search_window_xpath = "/html/body/div[1]/main/div[2]/div[3]/div[1]/input"
            search_window = driver.find_element(By.XPATH, search_window_xpath)
            search_window.send_keys(keyword)
            search_window.send_keys(Keys.ENTER)

            time.sleep(0.3)
            synonym_list_xpath = "/html/body/div[1]/main/div[2]/div[3]/button/span"
            driver.find_element(By.XPATH, synonym_list_xpath).click()

            # 유사어 입력
            for synonym in keywords[keyword]:
                synonym_input_xpath = f"/html/body/div[1]/main/div[2]/div[3]/div[2]/div[3]/div[{idx - adjust_idx_value}]/div/input"
                synonym_input_element = driver.find_element(By.XPATH, synonym_input_xpath)
                synonym_input_element.send_keys(synonym)
                synonym_input_element.send_keys(",")

            if (idx % 9 == 0) or keyword == list(keywords.keys())[-1]:  # 9개 or 마지막 | 10개 하면 연관검색어 넣을때 웹사이트 버그있음.
                synonym_confirm_xpath = "/html/body/div[1]/main/div[2]/div[3]/div[2]/button"
                driver.find_element(By.XPATH, synonym_confirm_xpath).click()

                keyword_analysis_btn_xpath = "/html/body/div[1]/main/div[2]/button"
                driver.find_element(By.XPATH, keyword_analysis_btn_xpath).click()

                # 일별 보기 클릭
                daily_setting_btn_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[1]/button[2]"
                selenium_execute('xpath', daily_setting_btn_xpath)
                time.sleep(5)

                # 캘린더
                open_set_yesterday_option_btn_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/input"
                selenium_execute('xpath', open_set_yesterday_option_btn_xpath, 5)

                # 전체데이터 다운을 위한 날짜 선택
                select_year_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[2]/button"
                select_2016_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[2]/ul/li[1]"
                selenium_execute("xpath", select_year_xpath)
                selenium_execute("xpath", select_2016_xpath)

                select_month_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[1]/button"
                select_01_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/thead/tr[1]/th[2]/div/div[1]/ul/li[1]"
                selenium_execute("xpath", select_month_xpath)
                selenium_execute("xpath", select_01_xpath)

                select_16_06_01_xpath = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[1]/table/tbody/tr[1]/td[6]/button"
                selenium_execute("xpath", select_16_06_01_xpath)

                col, row = get_calendar(date[:4], date[4:6], date[6:])
                select_today_xpath = f"/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[1]/div[2]/table/tbody/tr[{col}]/td[{row}]/button"
                selenium_execute("xpath", select_today_xpath)

                # 날짜 선택 후 적용 버튼
                temp_selectbtn = "/html/body/div[1]/main/div[2]/div[2]/div[2]/div/div/div[2]/button[2]"
                selenium_execute('xpath', temp_selectbtn, 2)
                time.sleep(20)

                # 파일 다운로드
                driver.execute_script('document.querySelector("#simple-result > div.f-c.container.rel > div > div.f-r.f-a-c.f-j-sb > div:nth-child(2) > button").click();')
                time.sleep(30)

                if not os.path.isfile(os.path.join(downloads_path, "Keyword_간편분석.xlsx")):  # 다운로드 되는 파일명
                    print("file download retry")
                    driver.execute_script('document.querySelector("#simple-result > div.f-c.container.rel > div > div.f-r.f-a-c.f-j-sb > div:nth-child(2) > button").click();')
                    time.sleep(20)

                keyword_1, keyword_2, keyword_3, keyword_4, keyword_5, keyword_6, keyword_7, keyword_8, keyword_9 = '', '', '', '', '', '', '', '', '',
                if os.path.isfile(os.path.join(downloads_path, "Keyword_간편분석.xlsx")):  # 다운로드 되는 파일명
                    print(str(keyword), " 파일 확인")
                    workbook = load_workbook(filename=os.path.join(downloads_path, "Keyword_간편분석.xlsx"))
                    sheet = workbook.active
                    last_column_value = len(sheet["A"])
                    date_column = [obj.value for obj in list(itertools.chain(*sheet["A8": f"A{last_column_value}"]))]

                    ## 단체 읽기 로직
                    keyword_1 = sheet["B6"].value
                    keyword_1_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["B8": f"B{last_column_value}"]))]
                    keyword_1_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["C8":f"C{last_column_value}"]))]
                    keyword_1_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["D8": f"D{last_column_value}"]))]

                    keyword_2 = sheet["E6"].value
                    keyword_2_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["E8": f"E{last_column_value}"]))]
                    keyword_2_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["F8":f"F{last_column_value}"]))]
                    keyword_2_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["G8": f"G{last_column_value}"]))]

                    keyword_3 = sheet["H6"].value
                    keyword_3_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["H8": f"H{last_column_value}"]))]
                    keyword_3_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["I8":f"I{last_column_value}"]))]
                    keyword_3_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["J8": f"J{last_column_value}"]))]

                    keyword_4 = sheet["K6"].value
                    keyword_4_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["K8": f"K{last_column_value}"]))]
                    keyword_4_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["L8":f"L{last_column_value}"]))]
                    keyword_4_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["M8": f"M{last_column_value}"]))]

                    keyword_5 = sheet["N6"].value
                    keyword_5_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["N8": f"N{last_column_value}"]))]
                    keyword_5_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["O8":f"O{last_column_value}"]))]
                    keyword_5_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["P8": f"P{last_column_value}"]))]

                    keyword_6 = sheet["Q6"].value
                    keyword_6_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["Q8": f"Q{last_column_value}"]))]
                    keyword_6_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["R8":f"R{last_column_value}"]))]
                    keyword_6_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["S8": f"S{last_column_value}"]))]

                    keyword_7 = sheet["T6"].value
                    keyword_7_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["T8": f"T{last_column_value}"]))]
                    keyword_7_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["U8":f"U{last_column_value}"]))]
                    keyword_7_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["V8": f"V{last_column_value}"]))]

                    keyword_8 = sheet["W6"].value
                    keyword_8_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["W8": f"W{last_column_value}"]))]
                    keyword_8_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["X8":f"X{last_column_value}"]))]
                    keyword_8_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["Y8": f"Y{last_column_value}"]))]

                    keyword_9 = sheet["Z6"].value
                    keyword_9_total_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["Z8": f"Z{last_column_value}"]))]
                    keyword_9_pc_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["AA8":f"AA{last_column_value}"]))]
                    keyword_9_mob_cnt = [obj.value if obj.value != '-' else 0 for obj in list(itertools.chain(*sheet["AB8": f"AB{last_column_value}"]))]

                    excel_keywords = keyword_1, keyword_2, keyword_3, keyword_4, keyword_5, keyword_6, keyword_7, keyword_8, keyword_9

                    if keyword_1 not in temp_keyword_last_updated_date.keys():  # 엑셀의 키워드가 db에 없는 단어면,
                        temp_keyword_last_updated_date[keyword_1] = '2015-12-31'
                    if keyword_2 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_2] = '2015-12-31'
                    if keyword_3 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_3] = '2015-12-31'
                    if keyword_4 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_4] = '2015-12-31'
                    if keyword_5 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_5] = '2015-12-31'
                    if keyword_6 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_6] = '2015-12-31'
                    if keyword_7 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_7] = '2015-12-31'
                    if keyword_8 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_8] = '2015-12-31'
                    if keyword_9 not in temp_keyword_last_updated_date.keys():
                        temp_keyword_last_updated_date[keyword_9] = '2015-12-31'

                    keyword = sheet['B2'].value

                    desired_filename = f"{keyword}_검색분석_검색량_트렌드_{date}.xlsx"  # 저장할 파일명 셋팅

                    for excl_keyword in zip(excel_keywords):  # 초기화
                        if excl_keyword[0] is not None:
                            data[excl_keyword[0]] = {}

                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_1]:  # date_column 값이 키워드 업데이트 날짜보다 큰 경우에만 인서트한다.
                            data[keyword_1][maderi_date] = {}
                            data[keyword_1][maderi_date].update({
                                "total_cnt": keyword_1_total_cnt[idx2],
                                "pc_cnt": keyword_1_pc_cnt[idx2],
                                "mobile_cnt": keyword_1_mob_cnt[idx2]
                            })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_2]:
                            if keyword_2 is not None:
                                data[keyword_2][maderi_date] = {}
                                data[keyword_2][maderi_date].update({
                                    "total_cnt": keyword_2_total_cnt[idx2],
                                    "pc_cnt": keyword_2_pc_cnt[idx2],
                                    "mobile_cnt": keyword_2_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_3]:
                            if keyword_3 is not None:
                                data[keyword_3][maderi_date] = {}
                                data[keyword_3][maderi_date].update({
                                    "total_cnt": keyword_3_total_cnt[idx2],
                                    "pc_cnt": keyword_3_pc_cnt[idx2],
                                    "mobile_cnt": keyword_3_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_4]:
                            if keyword_4 is not None:
                                data[keyword_4][maderi_date] = {}
                                data[keyword_4][maderi_date].update({
                                    "total_cnt": keyword_4_total_cnt[idx2],
                                    "pc_cnt": keyword_4_pc_cnt[idx2],
                                    "mobile_cnt": keyword_4_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_5]:
                            if keyword_5 is not None:
                                data[keyword_5][maderi_date] = {}
                                data[keyword_5][maderi_date].update({
                                    "total_cnt": keyword_5_total_cnt[idx2],
                                    "pc_cnt": keyword_5_pc_cnt[idx2],
                                    "mobile_cnt": keyword_5_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_6]:
                            if keyword_6 is not None:
                                data[keyword_6][maderi_date] = {}
                                data[keyword_6][maderi_date].update({
                                    "total_cnt": keyword_6_total_cnt[idx2],
                                    "pc_cnt": keyword_6_pc_cnt[idx2],
                                    "mobile_cnt": keyword_6_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_7]:
                            if keyword_7 is not None:
                                data[keyword_7][maderi_date] = {}
                                data[keyword_7][maderi_date].update({
                                    "total_cnt": keyword_7_total_cnt[idx2],
                                    "pc_cnt": keyword_7_pc_cnt[idx2],
                                    "mobile_cnt": keyword_7_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_8]:
                            if keyword_8 is not None:
                                data[keyword_8][maderi_date] = {}
                                data[keyword_8][maderi_date].update({
                                    "total_cnt": keyword_8_total_cnt[idx2],
                                    "pc_cnt": keyword_8_pc_cnt[idx2],
                                    "mobile_cnt": keyword_8_mob_cnt[idx2]
                                })
                    for idx2, maderi_date in enumerate(date_column):
                        if maderi_date > temp_keyword_last_updated_date[keyword_9]:
                            if keyword_9 is not None:
                                data[keyword_9][maderi_date] = {}
                                data[keyword_9][maderi_date].update({
                                    "total_cnt": keyword_9_total_cnt[idx2],
                                    "pc_cnt": keyword_9_pc_cnt[idx2],
                                    "mobile_cnt": keyword_9_mob_cnt[idx2]
                                })

                    # 종합분석과 다르게 검색분석은 한 엑셀에 키워드가 9개 있어서 키워드별로 각각 엑셀파일로 나누는 로직
                    split_searching_analysis_data_and_upload_s3_and(keyword_1, temp_keyword_last_updated_date[keyword_1], keyword_1_total_cnt, keyword_1_pc_cnt, keyword_1_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_2, temp_keyword_last_updated_date[keyword_2], keyword_2_total_cnt, keyword_2_pc_cnt, keyword_2_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_3, temp_keyword_last_updated_date[keyword_3], keyword_3_total_cnt, keyword_3_pc_cnt, keyword_3_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_4, temp_keyword_last_updated_date[keyword_4], keyword_4_total_cnt, keyword_4_pc_cnt, keyword_4_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_5, temp_keyword_last_updated_date[keyword_5], keyword_5_total_cnt, keyword_5_pc_cnt, keyword_5_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_6, temp_keyword_last_updated_date[keyword_6], keyword_6_total_cnt, keyword_6_pc_cnt, keyword_6_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_7, temp_keyword_last_updated_date[keyword_7], keyword_7_total_cnt, keyword_7_pc_cnt, keyword_7_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_8, temp_keyword_last_updated_date[keyword_8], keyword_8_total_cnt, keyword_8_pc_cnt, keyword_8_mob_cnt)
                    split_searching_analysis_data_and_upload_s3_and(keyword_9, temp_keyword_last_updated_date[keyword_9], keyword_9_total_cnt, keyword_9_pc_cnt, keyword_9_mob_cnt)

                    # s3에 저장
                    try:
                        os.system(f"rm {downloads_path}/*")
                    except Exception as e:
                        os.system(f"rm {downloads_path}/*")
                        driver.quit()
                        raise e

                else:
                    print(str(keyword) + " Trend File Not Exists")  # 키워드 나오도록 수정필요

                driver.get((url + "/keyword/simple"))
                adjust_idx_value += 9

                print("temp_keyword_last_updated_date value : ", temp_keyword_last_updated_date)

                try:
                    # DB 인서트
                    insert_query = f""""""

                    insert_sql_args = []
                    # db_Date : 최신업뎃날짜, keyword 는 사실 없어도댐. 왜냐면 여기는 여러개라서..
                    for _, db_date in temp_keyword_last_updated_date.items():
                        if keyword_1 in data.keys():
                            for maderi_date in data[keyword_1]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_1,
                                                            data[keyword_1][maderi_date]['total_cnt'],
                                                            data[keyword_1][maderi_date]['pc_cnt'],
                                                            data[keyword_1][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_1))
                        if keyword_2 in data.keys():
                            for maderi_date in data[keyword_2]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_2,
                                                            data[keyword_2][maderi_date]['total_cnt'],
                                                            data[keyword_2][maderi_date]['pc_cnt'],
                                                            data[keyword_2][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_2))
                        if keyword_3 in data.keys():
                            for maderi_date in data[keyword_3]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_3,
                                                            data[keyword_3][maderi_date]['total_cnt'],
                                                            data[keyword_3][maderi_date]['pc_cnt'],
                                                            data[keyword_3][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_3))
                        if keyword_4 in data.keys():
                            for maderi_date in data[keyword_4]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_4,
                                                            data[keyword_4][maderi_date]['total_cnt'],
                                                            data[keyword_4][maderi_date]['pc_cnt'],
                                                            data[keyword_4][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_4))
                        if keyword_5 in data.keys():
                            for maderi_date in data[keyword_5]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_5,
                                                            data[keyword_5][maderi_date]['total_cnt'],
                                                            data[keyword_5][maderi_date]['pc_cnt'],
                                                            data[keyword_5][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_5))
                        if keyword_6 in data.keys():
                            for maderi_date in data[keyword_6]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_6,
                                                            data[keyword_6][maderi_date]['total_cnt'],
                                                            data[keyword_6][maderi_date]['pc_cnt'],
                                                            data[keyword_6][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_6))
                        if keyword_7 in data.keys():
                            for maderi_date in data[keyword_7]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_7,
                                                            data[keyword_7][maderi_date]['total_cnt'],
                                                            data[keyword_7][maderi_date]['pc_cnt'],
                                                            data[keyword_7][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_7))
                        if keyword_8 in data.keys():
                            for maderi_date in data[keyword_8]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_8,
                                                            data[keyword_8][maderi_date]['total_cnt'],
                                                            data[keyword_8][maderi_date]['pc_cnt'],
                                                            data[keyword_8][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_8))
                        if keyword_9 in data.keys():
                            for maderi_date in data[keyword_9]:
                                if maderi_date > db_date:
                                    insert_sql_args.append((maderi_date,
                                                            keyword_9,
                                                            data[keyword_9][maderi_date]['total_cnt'],
                                                            data[keyword_9][maderi_date]['pc_cnt'],
                                                            data[keyword_9][maderi_date]['mobile_cnt']))
                                                           # maderi_date,
                                                           # keyword_9))


                    insert_sql_args = list(set(insert_sql_args))  # 중복제거
                    insert_sql_args.sort()
                    print("insert_sql_args : ", insert_sql_args)
                    data = dict(sorted(data.items(), reverse=True))
                    print("data : ", data)

                    util.query_ods("insert", insert_query, insert_sql_args)
                except Exception as e:
                    print("DB insert Error")
                    driver.quit()
                    raise e

            print("keyword_cnt : ", idx)
        driver.quit()

        # 다음 스케줄 동작 시 s3 업로드를 위한 파일 정리
        os.system(f"rm {downloads_path}/*")

    @task
    def related_keyword_total_data():
        # date = '20230710'
        maderi_bucket = "s3://edl-prod-archivei-bucket/maderi"
        url = 'https://www.maderi.co.kr'  # 홈 ( 로그인 )

        # 경로 설정
        cat = "related_keyword"
        present_path = os.getcwd()
        downloads_path = os.path.join(present_path, "data", cat)
        print(present_path)
        print(downloads_path)
        os.system(f"rm {downloads_path}/*")

        # 드라이버 옵션 설정
        chrome_options = Options()
        prefs = {
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }

        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option("prefs", {"download.default_directory": f"{downloads_path}"})
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")  # window 10, chrome

        chrome_options.add_argument("--headless")  # 서버용
        chrome_options.add_argument("--no-sandbox")  # 서버용

        driver_path = ChromeDriverManager().install()  # 드라이버 설치
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_window_size(1920, 1080)
        wait = WebDriverWait(driver, 60)

        bIstest = False

        if bIstest:
            import info
            pass
        else:
            chrome_options.add_argument("--headless")  # 서버용
            chrome_options.add_argument("--no-sandbox")  # 서버용
            import module.util as util
            from module.slack_alert import SlackAlert

        date = util.get_date().replace("-", "")  # '어제' 데이터가 네이버에 의존하며, 업데이트 시간이 오전 10~12시

        def measure_time(original_function):
            def wrapper(*args, **kwargs):
                start_time = time.time()
                result = original_function(*args, **kwargs)

                end_time = time.time()
                execution_time = end_time - start_time
                print(f"{original_function.__name__} 실행 시간: {execution_time}초")
                return result

            return wrapper

        @measure_time
        def selenium_execute(target, path, sec=0):
            print(path)
            try:
                if target == 'css':
                    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                    element.click()

                elif target == 'xpath':
                    element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                    element.click()

            except ElementClickInterceptedException:
                print("ElementClickInterceptedException error occurred, [Retrying..]")
                # driver.refresh()
                try:
                    if target == 'css':
                        element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                        element.click()

                    elif target == 'xpath':
                        element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                        element.click()

                except Exception as e:
                    err_msg = "selenium_execute ElementClickInterceptedException Error : " + str(e)
                    print(err_msg)
                    # raise err_msg
            except TimeoutException:
                print("TimeoutException error occurred, [Retrying..]")
                # driver.refresh()
                try:
                    if target == 'css':
                        element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path)))
                        element.click()

                    elif target == 'xpath':
                        element = wait.until(EC.presence_of_element_located((By.XPATH, path)))
                        element.click()
                except Exception as e:
                    err_msg = "selenium_execute TimeoutException Error : " + str(e)
                    raise err_msg

        def download_keyword_list():
            # 키워드 목록 파일 다운로드
            keyword_excel_download_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[2]/div[1]/div[2]/button[3]"
            selenium_execute("xpath", keyword_excel_download_btn_xpath, 4)
            time.sleep(7)  # 파일 다운로드 대기
            print("키워드 리스트 파일 다운로드 완료")

            desired_filename = f"Keyword_키워드 종합 진단.xlsx"  # 저장할 파일명 셋팅

            # 전체 키워드 가져오기 및 s3 업로드
            if os.path.isfile(os.path.join(downloads_path, desired_filename)):
                workbook = load_workbook(filename=os.path.join(downloads_path, desired_filename))
                sheet = workbook.active

                keyword_list = []
                start_cell = 'C7'
                end_cell = 'C106'
                for row in sheet[start_cell:end_cell]:
                    for cell in row:
                        keyword_list.append(cell.value)

                keyword_list = list(filter(None, keyword_list))
                keyword_cnt = len(keyword_list)  # None 제거 후 개수 카운트
                print(keyword_list)

                os.system(f"rm {downloads_path}/*")
                print("키워드 파일 삭제완료")

                time.sleep(10)
            else:
                raise Exception("Keyword_File Is Not Exist")

            return keyword_list

        def maderi_login(web_driver):
            # 마대리 접속
            web_driver.get(url + "/login")

            # 로그인 정보 입력 및 로그인
            id_xpath = "/html/body/div[1]/section/div[1]/div/div/form/div[1]/div/input"
            passwd_xpath = "/html/body/div[1]/section/div[1]/div/div/form/div[2]/div/input"
            login_btn_xpath = "/html/body/div[1]/section/div[1]/div/div/form/button"

            if bIstest:
                web_driver.find_element(By.ID, 'email').send_keys(info.userInfo2['iD'])
                web_driver.find_element(By.ID, 'password').send_keys(info.userInfo2['pW'])
            else:
                web_driver.find_element(By.ID, 'email').send_keys(maderi_id)
                web_driver.find_element(By.ID, 'password').send_keys(maderi_pw)

            selenium_execute('xpath', login_btn_xpath)

            # 로그인 대기
            time.sleep(30)
            print("로그인 끝")

        def control_aws_s3(aws_act, act_msg, start, finish=None):
            try:
                if finish is None:
                    s3_command = f"aws s3 {aws_act} '{start}'"
                else:
                    s3_command = f"aws s3 {aws_act} '{start}' {finish}"

                result = subprocess.run(s3_command, shell=True, capture_output=True, text=True)

                if result.returncode == 0:
                    print(f'{s3_filename} 파일 {act_msg} 성공')
                    print('표준 출력:', result.stdout)
                else:
                    print(f'{s3_filename} 파일 {act_msg} 실패')
                    print('에러 출력:', result.stderr)  #

            except Exception as e:
                os.system(f"rm {downloads_path}/*")
                driver.quit()
                raise e
        try:
            # 새 엑셀에 적용할 셀 효과
            fill = PatternFill(fill_type='solid', fgColor='DDEBF7')
            border_style = Side(border_style="thin", color="000000")  # 얇은 실선, 검은색
            border = Border(top=border_style, right=border_style, bottom=border_style, left=border_style)

            # 인서트 기록이 있는 키워드의 최근 날짜 가져오기
            sql = """"""
            keyword_last_updated_date = {} # keyword_last_updated_date = {'수학학원': "2023-07-17"}
            temp_keyword_last_updated_date = util.query_ods("select", sql)
            temp_keyword_last_updated_date = [[str(values[0]), values[1]] for values in temp_keyword_last_updated_date]
            for row in temp_keyword_last_updated_date:  # [0] : 업데이트 날짜, [1] : 키워드
                key = row[1]
                value = row[0].replace(" 00:00:00", "")
                keyword_last_updated_date[key] = value

            # 출력확인
            pprint("keyword_last_updated_date value : ")
            pprint(keyword_last_updated_date)

            maderi_login(driver)
            driver.get(url + "/keyword")

            keyword_list = download_keyword_list()
            keyword_cnt = len(keyword_list)

            ####
            # 순차적 검색
            start_keyword = 1  # 오류 발생 시, 출력된 keyword_cnt : 값을 넣으면 실패한 지점부터 실행
            last_cnt = start_keyword - 1
            for cnt in range(start_keyword, keyword_cnt + 1):
                # 현재 목록이 100개로 설정된 경우 스킵
                time.sleep(5)
                cnt_xpath = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/input"
                value = driver.find_element(By.XPATH, cnt_xpath).get_attribute("value")
                if value != '100개':
                    # 100개 보기
                    open_page_option_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/input"
                    driver.find_element(By.XPATH, open_page_option_btn_xpath).click()

                    # 100개 보기 선택
                    option_100 = "/html/body/div[1]/main/div[4]/article/div/div[1]/div/div[1]/div/div/ul/li[4]"
                    selenium_execute('xpath', option_100, 3)

                    # print(data) # 데이터 확인용
                print("keyword_cnt : ", cnt)

                # 분석할 키워드 선택
                select_box_xpath = f"/html/body/div[1]/main/div[4]/article/div/div[2]/div[2]/table/tbody/tr[{cnt}]/td[1]/div/label"
                selenium_execute('xpath', select_box_xpath)

                if (cnt % 10 == 0) or (cnt == keyword_cnt):  # 10개 or 마지막
                    # 검색분석 클릭
                    analysis_btn_xpath = "/html/body/div[1]/main/div[4]/article/div/div[3]/button[2]"
                    selenium_execute('xpath', analysis_btn_xpath, 10)

                    # 선택된 키워드 돌리기
                    for idx, _ in enumerate(keyword_list[last_cnt:cnt], start=1):
                        # 다음 키워드 선택
                        time.sleep(6)  # wait 'keyword_xpath' loading
                        keyword_xpath = f"/html/body/div[1]/main/div[4]/div[{idx}]"
                        selenium_execute('xpath', keyword_xpath, 10)
                        time.sleep(3)

                        try:
                            move_dw_btn = driver.find_element(By.XPATH, "/html/body/div[1]/main/div[6]/div[4]")
                            # 요소가 있으면 해당 위치로 스크롤 이동
                            driver.execute_script("arguments[0].scrollIntoView();", move_dw_btn)
                        except NoSuchElementException:
                            print("요소를 찾지 못했습니다.")

                        # 파일 다운로드
                        file_download_btn_xpath = "/html/body/div[1]/main/div[7]/div[1]/div[2]/button"
                        selenium_execute("xpath", file_download_btn_xpath)
                        time.sleep(7)

                        # 다운로드한 파일 읽고 변수에 할당
                        new_workbook = load_workbook(filename=os.path.join(downloads_path, "Keyword_검색분석_연관검색어.xlsx"))
                        new_sheet = new_workbook.active

                        keyword = new_sheet["B2"].value
                        day_cnt = len(new_sheet['6'])
                        is_null = new_sheet["A7"].value

                        print(is_null)
                        print(type(is_null))

                        # new 엑셀 데이터 읽기
                        data = {}
                        data[keyword] = {}
                        insert_sql_args = []

                        # data = {keyword: {}}
                        if is_null == 1:  # 데이터가 있는 경우
                            for idx in range(1, day_cnt):
                                for row in new_sheet.iter_rows(min_row=6, max_row=16):
                                    if row[idx].value.replace("-", "").isdigit():
                                        excel_date = row[idx].value
                                        data[keyword][excel_date] = {}
                                    else:
                                        data[keyword][excel_date][row[idx].value] = row[0].value
                        else:
                            driver.execute_script("window.scrollTo(0, 0);")
                            print(f"{keyword} 내용없음")
                            continue
                        print(data)

                        # 파일 삭제
                        os.system(f"rm {downloads_path}/*")

                        # 인서트 기록이 있는 단어인 경우
                        if keyword in keyword_last_updated_date:
                            # 기존 파일 다운받기
                            s3_filename = f'{keyword}_검색분석_연관검색어_{keyword_last_updated_date[keyword]}.xlsx'
                            file_downloaded_from_s3 = f'{keyword}_검색분석_연관검색어.xlsx'

                            start = f'{maderi_bucket}/연관검색어/{s3_filename}'
                            fin = f'{downloads_path}/{file_downloaded_from_s3}'
                            control_aws_s3('cp', '다운로드', start, fin)

                            # 읽은 데이터 쓰기
                            workbook = load_workbook(filename=os.path.join(downloads_path, file_downloaded_from_s3))
                            sheet = workbook.active

                            sheet_row_date = len(sheet["A"])
                            sheet_latest_date = sheet["A"][-1].value

                            # 이전 날짜 데이터 제거, 아래에 num 변수로 row 를 맞추기 위해 사전 삭제함
                            for date in sorted(data[keyword]):
                                if date <= str(sheet_latest_date):
                                    del data[keyword][date]

                            # 기존 엑셀에 추가하는 로직
                            for num, date in enumerate(sorted(data[keyword]), start=1):
                                if date > str(sheet_latest_date):
                                    # print(date, data[keyword][date])

                                    for keyword_related in data[keyword][date]:
                                        sheet[f'A{sheet_row_date + num}'] = date
                                        sheet[f"A{sheet_row_date + num}"].fill = fill
                                        sheet[f"A{sheet_row_date + num}"].border = border
                                        sheet[f"A{sheet_row_date + num}"].font = Font(bold=True)

                                        sheet.cell(row=sheet_row_date + num, column=data[keyword][date][keyword_related] + 1, value=keyword_related)

                        # 처음 인서트 하는 단어인 경우
                        else:
                            workbook = Workbook()
                            sheet = workbook.active

                            sheet_row_date = len(sheet["A"])
                            sheet_latest_date = '2016-01-01'
                            keyword_last_updated_date[keyword] = '2015-12-31'

                            # 새 엑셀 파일을 만드는 로직
                            for keyword in data:
                                # 엑셀 파일에 데이터 쌓기

                                # 헤더
                                sheet["A2"].font = Font(bold=True)
                                sheet["A2"] = "분석 대상"
                                sheet["B2"] = keyword

                                # 날짜, 순위
                                sheet["A6"].border = border
                                sheet["A6"].font = Font(bold=True)
                                sheet["A6"] = r"날짜\순위"

                                # B6 ~ K6 = 1 ~ 10 입력 및 색상 적용
                                for i, column in enumerate(
                                        sheet.iter_cols(min_col=2, max_col=11, min_row=6, max_row=6)):
                                    for j, cell in enumerate(column):
                                        cell.value = i + 1
                                        cell.border = border
                                        cell.fill = fill

                                # 날짜 쓰기
                                # 키워드 반복문
                                for idx_2, date in enumerate(sorted(data[keyword])):  # 오름차순 쓰기를 위한 sorted()

                                    # 새로운 키워드인 경우
                                    if sheet_latest_date is None:
                                        sheet[f"A{idx_2 + 7}"].fill = fill
                                        sheet[f"A{idx_2 + 7}"].border = border
                                        sheet[f"A{idx_2 + 7}"].font = Font(bold=True)
                                        sheet[f"A{idx_2 + 7}"] = date

                                        # 연관어 반복문
                                        for keyword_related in data[keyword][date]:
                                            sheet.cell(row=idx_2 + 7, column=data[keyword][date][keyword_related] + 1,
                                                       value=keyword_related)

                                    elif date > sheet_latest_date:  # 최신 날짜 데이트 인 경우 , 새로운 키워드인 경우 None이 나와서 str로 처리
                                        sheet[f"A{sheet_row_date + idx_2}"].fill = fill
                                        sheet[f"A{sheet_row_date + idx_2}"].border = border
                                        sheet[f"A{sheet_row_date + idx_2}"].font = Font(bold=True)
                                        sheet[f"A{sheet_row_date + idx_2}"] = date

                                        # 연관어 반복문
                                        for keyword_related in data[keyword][date]:
                                            sheet.cell(row=sheet_row_date + idx_2,
                                                       column=data[keyword][date][keyword_related] + 1,
                                                       value=keyword_related)

                        filename = f'{keyword}_검색분석_연관검색어.xlsx'
                        workbook.save(downloads_path + "/" + filename)

                        s3_filename = f'{keyword}_검색분석_연관검색어_{date}.xlsx'
                        delete_filename = f'{keyword}_검색분석_연관검색어_{keyword_last_updated_date[keyword]}.xlsx'
                        print("delete_filename value : ", delete_filename)

                        start = f"{maderi_bucket}/연관검색어/{delete_filename}"
                        control_aws_s3("rm", "삭제", start)

                        start = f"{downloads_path}/{filename}"
                        fin = f"{maderi_bucket}/연관검색어/{s3_filename}"
                        control_aws_s3("cp", "업로드", start, fin)


                        # 다음 키워드 클릭을 위한 스크롤 업
                        driver.execute_script("window.scrollTo(0, 0);")

                        # 인서트 만들기
                        insert_query = f""""""
                        for date in data[keyword]:
                            for ii, related_keyword in enumerate(data[keyword][date]):
                                if '-' == list(data[keyword][date].keys())[ii]:
                                    insert_sql_args.append((date,
                                                            keyword,
                                                            '-',
                                                            0,
                                                            ))

                                elif date > sheet_latest_date:
                                    insert_sql_args.append((date,
                                                            keyword,
                                                            related_keyword,
                                                            data[keyword][date][related_keyword],
                                                            ))

                        insert_sql_args = list(set(insert_sql_args))
                        print(sorted(insert_sql_args))

                    # 인서트
                        util.query_ods("insert", insert_query, insert_sql_args)

                    last_cnt = cnt
                    driver.get(url + "/keyword")

        except Exception as e:
            print("Collecting Error", str(e))
            driver.quit()
            raise e
        driver.quit()
        os.system(f"rm {downloads_path}/*")

    [registered_keyword_total_data() >> searched_keyword_total_data() >> related_keyword_total_data()]

dag = maderi_crawling()
